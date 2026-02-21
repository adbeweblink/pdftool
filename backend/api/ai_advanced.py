"""
AI 進階功能 API - 合約比對、個資偵測、表格提取、智能重命名
"""
import fitz
import base64
import json
import re
import httpx
from io import BytesIO
from fastapi import APIRouter, UploadFile, File, HTTPException, Form
from fastapi.responses import JSONResponse, StreamingResponse
from typing import List, Optional
from pydantic import BaseModel

from utils.file_handler import save_upload_file, generate_output_path

router = APIRouter()

# Gemini API 設定
GEMINI_API_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"


# ============ 工具函式 ============
def extract_pdf_text(file_path, max_pages: int = 50) -> str:
    """提取 PDF 全部文字"""
    pdf = fitz.open(file_path)
    text = ""

    for i, page in enumerate(pdf):
        if i >= max_pages:
            break
        text += f"\n=== 第 {i + 1} 頁 ===\n"
        text += page.get_text()

    pdf.close()
    return text


def extract_pdf_tables(file_path) -> list:
    """提取 PDF 中的表格結構"""
    pdf = fitz.open(file_path)
    all_tables = []

    for page_num, page in enumerate(pdf):
        # 使用 PyMuPDF 的表格提取功能
        tables = page.find_tables()

        for table_num, table in enumerate(tables):
            table_data = {
                "page": page_num + 1,
                "table_index": table_num + 1,
                "rows": []
            }

            # 提取表格內容
            for row in table.extract():
                table_data["rows"].append(row)

            if table_data["rows"]:
                all_tables.append(table_data)

    pdf.close()
    return all_tables


async def call_gemini_advanced(prompt: str, api_key: str, max_tokens: int = 4096) -> str:
    """呼叫 Gemini API（進階版，支援更長輸出）- BYOK 模式"""
    if not api_key:
        raise HTTPException(
            status_code=400,
            detail="⚠️ 請提供您的 Gemini API Key 才能使用此功能。"
        )

    async with httpx.AsyncClient(timeout=120.0) as client:
        try:
            response = await client.post(
                f"{GEMINI_API_URL}?key={api_key}",
                json={
                    "contents": [{
                        "parts": [{"text": prompt}]
                    }],
                    "generationConfig": {
                        "temperature": 0.3,  # 低溫度確保精確性
                        "maxOutputTokens": max_tokens,
                    }
                }
            )

            if response.status_code == 429:
                raise HTTPException(
                    status_code=429,
                    detail="⏳ AI 服務暫時繁忙，請稍等幾分鐘後再試。"
                )

            if response.status_code != 200:
                raise HTTPException(
                    status_code=response.status_code,
                    detail=f"❌ AI 服務發生錯誤，請稍後重試。錯誤代碼：{response.status_code}"
                )

            result = response.json()

            try:
                return result["candidates"][0]["content"]["parts"][0]["text"]
            except (KeyError, IndexError):
                raise HTTPException(
                    status_code=500,
                    detail="❌ 無法解析 AI 回應，請重試。"
                )

        except httpx.TimeoutException:
            raise HTTPException(
                status_code=504,
                detail="⏱️ AI 處理超時，請嘗試較小的文件或稍後重試。"
            )


# ============ 合約比對 ============
@router.post("/compare")
async def compare_contracts(
    file1: UploadFile = File(..., description="第一份合約 PDF"),
    file2: UploadFile = File(..., description="第二份合約 PDF"),
    focus_areas: str = Form(
        "all",
        description="關注重點：all=全部, terms=條款, numbers=金額日期, parties=當事人"
    ),
    api_key: str = Form(..., description="您的 Gemini API Key")
):
    """
    比對兩份合約的差異

    - 上傳兩份 PDF 合約
    - AI 分析並列出所有差異
    - 標示重要變更（金額、日期、責任條款等）
    """
    file1_path = await save_upload_file(file1, "ai")
    file2_path = await save_upload_file(file2, "ai")

    try:
        # 提取兩份文件的文字
        text1 = extract_pdf_text(file1_path)
        text2 = extract_pdf_text(file2_path)

        # 建立比對提示
        focus_prompts = {
            "all": "所有差異",
            "terms": "條款和條件的變更",
            "numbers": "金額、日期、數量的變更",
            "parties": "當事人、簽署者、聯絡資訊的變更"
        }
        focus = focus_prompts.get(focus_areas, focus_prompts["all"])

        prompt = f"""你是專業的合約審查專家。請仔細比對以下兩份合約，找出{focus}。

## 合約 A（原版）：
{text1[:15000]}

## 合約 B（新版）：
{text2[:15000]}

## 請以以下格式輸出：

### 📊 差異總覽
- 發現 X 處差異
- 重大變更：X 處
- 一般變更：X 處

### 🔴 重大差異（需特別注意）
對於每個重大差異，請列出：
1. **位置**：第 X 條 / 第 X 頁
2. **原版內容**：「...」
3. **新版內容**：「...」
4. **影響評估**：這個變更可能帶來的影響

### 🟡 一般差異
列出其他較小的差異

### 💡 審查建議
基於差異分析，給出具體建議

請用繁體中文回答，確保輸出清晰易讀。
"""

        result = await call_gemini_advanced(prompt, api_key)

        return {
            "success": True,
            "comparison": result,
            "file1_name": file1.filename,
            "file2_name": file2.filename,
            "focus_areas": focus_areas
        }

    finally:
        file1_path.unlink(missing_ok=True)
        file2_path.unlink(missing_ok=True)


# ============ 個資偵測 ============
@router.post("/pii-detect")
async def detect_pii(
    file: UploadFile = File(...),
    action: str = Form("detect", description="動作：detect=僅偵測, redact=偵測並遮蔽"),
    pii_types: str = Form(
        "all",
        description="個資類型：all, id=身分證, phone=電話, email=電子郵件, address=地址, account=帳號"
    ),
    api_key: str = Form(..., description="您的 Gemini API Key")
):
    """
    偵測並遮蔽 PDF 中的個人資料

    - 支援身分證字號、電話、Email、地址、銀行帳號
    - 可選擇僅偵測或直接遮蔽
    """
    file_path = await save_upload_file(file, "ai")

    try:
        text = extract_pdf_text(file_path)

        # 定義各類個資的正則表達式
        pii_patterns = {
            "id": {
                "name": "身分證字號",
                "patterns": [
                    r'[A-Z][12]\d{8}',  # 台灣身分證
                    r'[A-Z][89]\d{8}',  # 外籍居留證
                ]
            },
            "phone": {
                "name": "電話號碼",
                "patterns": [
                    r'09\d{2}[-\s]?\d{3}[-\s]?\d{3}',  # 手機
                    r'0\d[-\s]?\d{3,4}[-\s]?\d{4}',    # 市話
                    r'\+886[-\s]?\d{1,2}[-\s]?\d{3,4}[-\s]?\d{4}',  # 國際格式
                ]
            },
            "email": {
                "name": "電子郵件",
                "patterns": [
                    r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',
                ]
            },
            "address": {
                "name": "地址",
                "patterns": [
                    r'[\u4e00-\u9fff]+[縣市][\u4e00-\u9fff]+[區鄉鎮市][\u4e00-\u9fff]+[路街巷弄號樓室]+[\d\-之\u4e00-\u9fff]*',
                ]
            },
            "account": {
                "name": "銀行帳號",
                "patterns": [
                    r'\d{3,4}[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{2,4}',  # 銀行帳號格式
                ]
            }
        }

        # 選擇要偵測的類型
        if pii_types == "all":
            selected_types = pii_patterns.keys()
        else:
            selected_types = [t.strip() for t in pii_types.split(",")]

        # 執行偵測
        found_pii = []
        for pii_type in selected_types:
            if pii_type in pii_patterns:
                info = pii_patterns[pii_type]
                for pattern in info["patterns"]:
                    matches = re.findall(pattern, text)
                    for match in matches:
                        found_pii.append({
                            "type": pii_type,
                            "type_name": info["name"],
                            "value": match,
                            "masked": mask_pii(match, pii_type)
                        })

        # 使用 AI 補充偵測
        if found_pii or True:  # 總是使用 AI 補充
            prompt = f"""請分析以下文件，找出所有可能的個人資料（PII）。

文件內容：
{text[:10000]}

請找出：
1. 姓名
2. 身分證字號
3. 電話號碼
4. 電子郵件
5. 地址
6. 銀行帳號
7. 信用卡號
8. 護照號碼
9. 其他敏感個資

請以 JSON 格式回傳，格式如下：
```json
{{
    "pii_items": [
        {{"type": "name", "value": "王小明", "location": "第1頁"}},
        {{"type": "phone", "value": "0912-345-678", "location": "第2頁"}}
    ],
    "risk_level": "high/medium/low",
    "summary": "簡短摘要"
}}
```
"""
            ai_result = await call_gemini_advanced(prompt, api_key)

            # 嘗試解析 AI 回傳的 JSON
            try:
                # 提取 JSON 部分
                json_match = re.search(r'```json\s*([\s\S]*?)\s*```', ai_result)
                if json_match:
                    ai_pii = json.loads(json_match.group(1))
                else:
                    ai_pii = json.loads(ai_result)
            except:
                ai_pii = {"pii_items": [], "summary": ai_result}

        # 如果需要遮蔽
        output_file = None
        if action == "redact" and found_pii:
            output_file = await redact_pii_in_pdf(file_path, found_pii)

        result = {
            "success": True,
            "detected_count": len(found_pii),
            "regex_detected": found_pii,
            "ai_detected": ai_pii.get("pii_items", []),
            "risk_level": ai_pii.get("risk_level", "unknown"),
            "summary": ai_pii.get("summary", ""),
            "action": action
        }

        if output_file:
            result["download_url"] = f"/download/{output_file.name}"

        return result

    finally:
        file_path.unlink(missing_ok=True)


def mask_pii(value: str, pii_type: str) -> str:
    """遮蔽個資"""
    if pii_type == "id":
        return value[0] + "*" * 7 + value[-2:]
    elif pii_type == "phone":
        return value[:4] + "***" + value[-3:]
    elif pii_type == "email":
        parts = value.split("@")
        return parts[0][:2] + "***@" + parts[1]
    elif pii_type == "account":
        return value[:4] + "****" + value[-4:]
    else:
        if len(value) > 4:
            return value[:2] + "*" * (len(value) - 4) + value[-2:]
        return "*" * len(value)


async def redact_pii_in_pdf(file_path, pii_items: list):
    """在 PDF 中遮蔽個資"""
    pdf = fitz.open(file_path)

    for page in pdf:
        for item in pii_items:
            value = item["value"]
            # 搜尋並遮蔽
            text_instances = page.search_for(value)
            for inst in text_instances:
                # 用黑色矩形遮蔽
                page.draw_rect(inst, color=(0, 0, 0), fill=(0, 0, 0))

    output_path = generate_output_path("redacted.pdf")
    pdf.save(str(output_path))
    pdf.close()

    return output_path


# ============ 表格提取 ============
@router.post("/extract-table")
async def extract_tables(
    file: UploadFile = File(...),
    output_format: str = Form("json", description="輸出格式：json, csv, excel"),
    use_ai: bool = Form(True, description="是否使用 AI 增強提取"),
    api_key: str = Form(None, description="您的 Gemini API Key（使用 AI 增強時必填）")
):
    """
    從 PDF 提取表格並轉換為結構化資料

    - 自動偵測 PDF 中的表格
    - 支援匯出為 JSON、CSV、Excel
    - AI 增強可處理複雜表格
    """
    file_path = await save_upload_file(file, "ai")

    try:
        # 先用 PyMuPDF 提取表格
        tables = extract_pdf_tables(file_path)

        # 如果啟用 AI 增強，使用 AI 改善表格結構
        if use_ai and tables:
            if not api_key:
                raise HTTPException(status_code=400, detail="⚠️ 使用 AI 增強功能需要提供您的 Gemini API Key")
            prompt = f"""請分析以下從 PDF 提取的表格資料，並：
1. 修正任何提取錯誤
2. 識別表頭
3. 統一資料格式
4. 合併跨頁表格（如果適用）

原始表格資料：
{json.dumps(tables[:5], ensure_ascii=False, indent=2)}

請以 JSON 格式回傳優化後的表格：
```json
{{
    "tables": [
        {{
            "title": "表格標題（如果有）",
            "headers": ["欄位1", "欄位2"],
            "rows": [
                ["值1", "值2"],
                ["值3", "值4"]
            ],
            "summary": "表格簡述"
        }}
    ]
}}
```
"""
            ai_result = await call_gemini_advanced(prompt, api_key)

            try:
                json_match = re.search(r'```json\s*([\s\S]*?)\s*```', ai_result)
                if json_match:
                    enhanced_tables = json.loads(json_match.group(1))
                    tables = enhanced_tables.get("tables", tables)
            except:
                pass  # 使用原始表格

        # 如果沒有偵測到表格，嘗試用 AI 從文字中提取
        if not tables and api_key:
            text = extract_pdf_text(file_path)
            prompt = f"""請從以下文件中提取所有表格資料。即使是用空格或 Tab 對齊的資料也算表格。

文件內容：
{text[:10000]}

請以 JSON 格式回傳：
```json
{{
    "tables": [
        {{
            "title": "表格標題",
            "headers": ["欄位1", "欄位2"],
            "rows": [["值1", "值2"]]
        }}
    ],
    "found_count": 1
}}
```

如果沒有找到任何表格，回傳空陣列。
"""
            ai_result = await call_gemini_advanced(prompt, api_key)

            try:
                json_match = re.search(r'```json\s*([\s\S]*?)\s*```', ai_result)
                if json_match:
                    ai_tables = json.loads(json_match.group(1))
                    tables = ai_tables.get("tables", [])
            except:
                tables = []

        # 根據輸出格式處理
        if output_format == "csv":
            # 轉為 CSV
            csv_content = ""
            for i, table in enumerate(tables):
                csv_content += f"# 表格 {i + 1}\n"
                if "headers" in table:
                    csv_content += ",".join(str(h) for h in table["headers"]) + "\n"
                for row in table.get("rows", []):
                    csv_content += ",".join(str(cell) if cell else "" for cell in row) + "\n"
                csv_content += "\n"

            return StreamingResponse(
                iter([csv_content]),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=tables.csv"}
            )

        elif output_format == "excel":
            # 轉為 Excel（需要 openpyxl）
            try:
                from openpyxl import Workbook

                wb = Workbook()

                for i, table in enumerate(tables):
                    if i == 0:
                        ws = wb.active
                        ws.title = f"表格{i + 1}"
                    else:
                        ws = wb.create_sheet(f"表格{i + 1}")

                    row_num = 1
                    if "headers" in table:
                        for col, header in enumerate(table["headers"], 1):
                            ws.cell(row=row_num, column=col, value=header)
                        row_num += 1

                    for row in table.get("rows", []):
                        for col, cell in enumerate(row, 1):
                            ws.cell(row=row_num, column=col, value=cell)
                        row_num += 1

                output = BytesIO()
                wb.save(output)
                output.seek(0)

                return StreamingResponse(
                    output,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=tables.xlsx"}
                )
            except ImportError:
                raise HTTPException(
                    status_code=500,
                    detail="❌ Excel 匯出需要安裝 openpyxl。請執行：pip install openpyxl"
                )

        # 預設 JSON 格式
        return {
            "success": True,
            "table_count": len(tables),
            "tables": tables,
            "source_file": file.filename
        }

    finally:
        file_path.unlink(missing_ok=True)


# ============ 智能重命名 ============
@router.post("/smart-rename")
async def smart_rename(
    file: UploadFile = File(...),
    naming_pattern: str = Form(
        "auto",
        description="命名模式：auto=自動, date_title=日期_標題, type_date=類型_日期"
    ),
    include_date: bool = Form(True, description="是否包含日期"),
    max_length: int = Form(50, description="檔名最大長度"),
    api_key: str = Form(..., description="您的 Gemini API Key")
):
    """
    根據 PDF 內容智能生成檔名

    - 分析文件內容
    - 提取關鍵資訊（日期、標題、類型）
    - 生成有意義的檔名
    """
    file_path = await save_upload_file(file, "ai")

    try:
        text = extract_pdf_text(file_path, max_pages=5)  # 只讀前 5 頁

        prompt = f"""請分析以下文件並提取關鍵資訊，用於生成檔案名稱。

文件內容：
{text[:5000]}

請提取以下資訊並以 JSON 格式回傳：
```json
{{
    "document_type": "合約/發票/報告/簡報/表單/信件/其他",
    "title": "文件標題或主題",
    "date": "YYYY-MM-DD 格式的日期（如果有）",
    "parties": ["相關單位或人名"],
    "key_identifier": "關鍵識別碼（如合約編號、發票號碼等）",
    "suggested_names": [
        "建議檔名1",
        "建議檔名2",
        "建議檔名3"
    ]
}}
```

命名規則：
1. 使用繁體中文
2. 不要使用特殊字元（\ / : * ? " < > |）
3. 長度不超過 {max_length} 字元
4. 要有辨識度，方便日後搜尋
"""

        ai_result = await call_gemini_advanced(prompt, api_key)

        try:
            json_match = re.search(r'```json\s*([\s\S]*?)\s*```', ai_result)
            if json_match:
                doc_info = json.loads(json_match.group(1))
            else:
                doc_info = json.loads(ai_result)
        except:
            doc_info = {
                "document_type": "文件",
                "title": "未知",
                "suggested_names": [file.filename]
            }

        # 根據命名模式生成檔名
        suggested_names = doc_info.get("suggested_names", [])

        if naming_pattern == "date_title":
            date = doc_info.get("date", "")
            title = doc_info.get("title", "文件")
            if date and include_date:
                primary_name = f"{date}_{title}"
            else:
                primary_name = title
        elif naming_pattern == "type_date":
            doc_type = doc_info.get("document_type", "文件")
            date = doc_info.get("date", "")
            if date and include_date:
                primary_name = f"{doc_type}_{date}"
            else:
                primary_name = doc_type
        else:
            # auto 模式使用 AI 建議
            primary_name = suggested_names[0] if suggested_names else "文件"

        # 清理檔名
        primary_name = re.sub(r'[\\/:*?"<>|]', '_', primary_name)
        primary_name = primary_name[:max_length]

        return {
            "success": True,
            "original_name": file.filename,
            "suggested_name": primary_name + ".pdf",
            "alternative_names": [n + ".pdf" for n in suggested_names[:3]],
            "document_info": doc_info,
            "naming_pattern": naming_pattern
        }

    finally:
        file_path.unlink(missing_ok=True)


# ============ 批次智能重命名 ============
@router.post("/batch-smart-rename")
async def batch_smart_rename(
    files: List[UploadFile] = File(...),
    naming_pattern: str = Form("auto"),
    api_key: str = Form(..., description="您的 Gemini API Key")
):
    """
    批次智能重命名多個 PDF
    """
    results = []

    for file in files:
        try:
            result = await smart_rename(
                file=file,
                naming_pattern=naming_pattern,
                api_key=api_key
            )
            results.append({
                "original": file.filename,
                "suggested": result["suggested_name"],
                "success": True
            })
        except Exception as e:
            results.append({
                "original": file.filename,
                "error": str(e),
                "success": False
            })

    return {
        "success": True,
        "total": len(files),
        "results": results
    }
